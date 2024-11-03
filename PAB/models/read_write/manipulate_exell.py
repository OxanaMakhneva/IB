import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from models.read_write.transport_in_BD import TransportModel

#Метод для конструирования таблицы в exell
def create_exell_table(workbook, sheet, headers, data):
    wsheet = workbook.create_sheet(sheet)
    # добавляем названия колонок
    wsheet.append(headers)
    # добавляем данные к таблице
    for row in data:
        wsheet.append(row)
    cust_table_width(workbook, wsheet)

#Метод для экспорта объекта ExportModel в exell
def create_myexell(sets, file_name):
    # создаем книгу
    workbook = Workbook()
    for set in sets:
        create_exell_table(workbook, set.data_lable, set.data_header, set.data_table)
    del workbook['Sheet']
    workbook.save(file_name)

#Метод для автоподстройки ширины колонок в exell
def cust_table_width(workbook, wsheet):
    # размер шрифта документа
    font_size = 11
    # словарь с размерами столбцов
    cols_dict = {}
    #определяем словарь длин по заголовку
    head = wsheet[1]
    for cell in head:
        # получаем букву текущего столбца
        letter = cell.column_letter
        # если в ячейке записаны данные
        if cell.value:
            # устанавливаем в ячейке размер шрифта
            cell.font = Font(name='Times New Roman', size = font_size)
            cell.alignment = Alignment(horizontal='center')
            # вычисляем количество символов, записанных в ячейку
            len_cell = len(str(cell.value))
            #записываем длину ячейки заголовка в словарь
            cols_dict[letter] = len_cell
    # проходимся по всем строкам документа кроме заголовка
    for idx, row in enumerate(wsheet.rows):
    # теперь по ячейкам каждой строки
        if idx > 0:
            for cell in row:
                letter = cell.column_letter
                if cell.value:
                    cell.font = Font(name='Times New Roman', size = font_size)
                    cell.alignment = Alignment(horizontal='center')
                    len_cell = len(str(cell.value))
                    if len_cell > cols_dict.get(letter, 0):
                        cols_dict[letter] = len_cell
                len_cell = cols_dict[letter]
                #расчитываем новую ширину столбца
                new_width_col = len_cell * font_size**(font_size*0.015)
                # применение новой ширины столбца
                wsheet.column_dimensions[cell.column_letter].width = new_width_col

def read_myexell_sheet(file_name):
    workbook = load_workbook(file_name)
    sheets = workbook.get_sheet_names()
    return sheets

def read_myexell_head(file_name, sheet_name, first_row):
    if not first_row:
        return None
    else:
        workbook = load_workbook(file_name)
        sheet = workbook.get_sheet_by_name(sheet_name)
        try:
            header = [cell.value for cell in sheet[first_row] if cell.value]
            clean_header = [re.sub('\n', ' ', head) for head in header]
            return clean_header
        except:
            return None

def read_myexell(file_name, sheet_name, first_row):
    workbook = load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    #определяем размер данных
    rows = [row for row in sheet.rows]
    header = [cell.value for cell in sheet[first_row] if cell.value]
    data_table = []
    for row_numb in range(first_row + 1, len(rows) + 1):
        record = [sheet.cell(row = row_numb, column = column_numb).value for column_numb in range(1, len(header) + 1)]
        data_table.append(record)
    data_set = TransportModel(sheet_name, header, data_table)
    return data_set

def color_fill(file_name):
    sheets = read_myexell_sheet(file_name)
    workbook = load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheets[0])
    for ind in range(1, sheet.max_row):
        accuracy = sheet.cell(row = ind, column = 1).value
        #print(accuracy)
        sheet.cell(row = ind, column = 1).fill = PatternFill(start_color="8a2be2", end_color="8a2be2", fill_type="solid")
        print(sheet.cell(row = ind, column = 1).fill)
        if "точно".find(accuracy):
            sheet.cell(row = ind, column = 1).fill = PatternFill(fill_type = 'solid', fgColor = "DDDDDD")
        elif "сомнительно".find(accuracy):
            sheet.cell(row = ind, column = 1).fill = PatternFill(fill_type = 'solid', fgColor = "FF0000")
        else:
            sheet.cell(row = ind, column = 1).fill = PatternFill(fill_type = 'solid', fgColor = "FFFFFF")
